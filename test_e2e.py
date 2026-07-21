"""End-to-end test including Excel export"""
import urllib.request
import json
import os
import sys


def test_full_workflow():
    print("=== Running End-to-End Test ===\n")

    # Initialize session
    print("Step 1: Create session...")
    init_data = json.dumps({'name': 'テスト三郎'}).encode('utf-8')
    request = urllib.request.Request(
        'http://localhost:5000/api/init',
        data=init_data,
        headers={'Content-Type': 'application/json'}
    )
    response = urllib.request.urlopen(request)
    session = json.loads(response.read().decode())
    session_id = session['session_id']
    total_files = session['total_files']
    print(
        f"   OK: Session {session_id[:8]}... created with {total_files} files\n")

    # Submit evaluations for first 10 files
    print(f"Step 2: Submit evaluations for first 10 files...")
    for i in range(10):
        eval_data = json.dumps({'rating': (i % 9) + 1}).encode('utf-8')
        request = urllib.request.Request(
            f'http://localhost:5000/api/submit-evaluation/{session_id}',
            data=eval_data,
            headers={'Content-Type': 'application/json'}
        )
        response = urllib.request.urlopen(request)
        result = json.loads(response.read().decode())
        print(f"   Submitted evaluation {i+1}/10 (rating: {(i % 9) + 1})")
    print()

    # Export results
    print("Step 3: Export results to Excel...")
    export_request = urllib.request.Request(
        f'http://localhost:5000/api/export-results/{session_id}'
    )
    export_request.get_method = lambda: 'POST'
    response = urllib.request.urlopen(export_request)
    result = json.loads(response.read().decode())

    if result['success']:
        filename = result['filename']
        filepath = result['filepath']
        print(f"   OK: Excel file created: {filename}")
        print(f"   OK: File saved at: {filepath}\n")

        # Check if file exists
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"Step 4: Verify exported file...")
            print(f"   OK: File exists")
            print(f"   OK: File size: {file_size} bytes")

            # Try to read with openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath)
                ws = wb.active
                row_count = ws.max_row
                col_count = ws.max_column
                print(
                    f"   OK: Excel sheet has {row_count} rows and {col_count} columns")
                print(
                    f"   OK: Data rows: {row_count - 1} (excluding header)\n")

                print("=== End-to-End Test PASSED ===")
                return True
            except Exception as e:
                print(f"   ERROR reading Excel: {e}\n")
        else:
            print(f"   ERROR: File not found at {filepath}\n")
    else:
        print(f"   ERROR: Export failed\n")

    return False


if __name__ == '__main__':
    try:
        success = test_full_workflow()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

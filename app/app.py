from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import os
import random
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import uuid

app = Flask(__name__)
CORS(app)

# Configuration
AUDIO_FOLDER = Path(__file__).parent.parent / "5_potechi_renamed"
RESULTS_FOLDER = Path(__file__).parent.parent / "results"
RESULTS_FOLDER.mkdir(exist_ok=True)

# Session storage (in-memory for simplicity, could use file/DB for persistence)
sessions = {}


def get_wav_files():
    """Get all WAV files from the audio folder"""
    if not AUDIO_FOLDER.exists():
        return []
    
    wav_files = sorted([
        f.stem for f in AUDIO_FOLDER.glob("*.wav")
    ])
    return wav_files


def create_randomized_list(wav_files):
    """Create a randomized list of wav files"""
    randomized = wav_files.copy()
    random.shuffle(randomized)
    return randomized


@app.route('/')
def index():
    """Serve the main page"""
    return render_template('index.html')


@app.route('/api/init', methods=['POST'])
def init_session():
    """Initialize a new session with participant name"""
    data = request.json
    participant_name = data.get('name', '').strip()
    
    if not participant_name:
        return jsonify({'error': 'Participant name is required'}), 400
    
    # Get all WAV files
    wav_files = get_wav_files()
    if not wav_files:
        return jsonify({'error': 'No WAV files found'}), 400
    
    # Create randomized list
    randomized_files = create_randomized_list(wav_files)
    
    # Create session
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        'participant_name': participant_name,
        'wav_files': randomized_files,
        'current_index': 0,
        'results': [],
        'current_repeat_count': 0,
    }
    
    return jsonify({
        'session_id': session_id,
        'total_files': len(randomized_files),
        'current_file': randomized_files[0]
    })


@app.route('/api/get-current-file/<session_id>', methods=['GET'])
def get_current_file(session_id):
    """Get current file information"""
    if session_id not in sessions:
        return jsonify({'error': 'Invalid session'}), 400
    
    session = sessions[session_id]
    current_index = session['current_index']
    wav_files = session['wav_files']
    
    if current_index >= len(wav_files):
        return jsonify({
            'finished': True,
            'total_files': len(wav_files),
            'current_index': current_index
        })
    
    return jsonify({
        'finished': False,
        'file_name': wav_files[current_index],
        'current_index': current_index + 1,
        'total_files': len(wav_files)
    })


@app.route('/api/audio/<session_id>/<filename>')
def get_audio(session_id, filename):
    """Stream audio file"""
    if session_id not in sessions:
        return jsonify({'error': 'Invalid session'}), 400
    
    audio_path = AUDIO_FOLDER / f"{filename}.wav"
    if not audio_path.exists():
        return jsonify({'error': 'Audio file not found'}), 404
    
    return send_file(str(audio_path), mimetype='audio/wav')


@app.route('/api/submit-evaluation/<session_id>', methods=['POST'])
def submit_evaluation(session_id):
    """Submit evaluation for current file"""
    if session_id not in sessions:
        return jsonify({'error': 'Invalid session'}), 400
    
    data = request.json
    rating = data.get('rating')
    
    # Validate rating
    if rating is not None and (rating < 1 or rating > 9):
        return jsonify({'error': 'Invalid rating'}), 400
    
    session = sessions[session_id]
    current_index = session['current_index']
    wav_files = session['wav_files']
    
    if current_index >= len(wav_files):
        return jsonify({'error': 'All files already evaluated'}), 400
    
    current_file = wav_files[current_index]
    
    # If rating is provided (not repeat)
    if rating is not None:
        result = {
            'file_name': current_file,
            'rating': rating,
            'order': current_index + 1,
            'repeat_count': session['current_repeat_count'],
            'timestamp': datetime.now().isoformat()
        }
        session['results'].append(result)
        session['current_index'] += 1
        session['current_repeat_count'] = 0
    else:
        # Repeat
        session['current_repeat_count'] += 1
    
    # Check if finished
    if session['current_index'] >= len(wav_files):
        return jsonify({
            'finished': True,
            'next_file': None,
            'current_index': session['current_index'],
            'total_files': len(wav_files)
        })
    
    return jsonify({
        'finished': False,
        'next_file': wav_files[session['current_index']],
        'current_index': session['current_index'] + 1,
        'total_files': len(wav_files)
    })


@app.route('/api/export-results/<session_id>', methods=['POST'])
def export_results(session_id):
    """Export evaluation results to Excel"""
    if session_id not in sessions:
        return jsonify({'error': 'Invalid session'}), 400
    
    session = sessions[session_id]
    participant_name = session['participant_name']
    results = session['results']
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "評価結果"
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    
    # Header row
    headers = ['ファイル名', '評価値', '再生順', 'リピート回数', '回答時刻']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1).value = result['file_name']
        ws.cell(row=row_idx, column=2).value = result['rating']
        ws.cell(row=row_idx, column=3).value = result['order']
        ws.cell(row=row_idx, column=4).value = result['repeat_count']
        
        # Format timestamp
        timestamp = datetime.fromisoformat(result['timestamp'])
        ws.cell(row=row_idx, column=5).value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        
        # Center align numeric columns
        for col in [2, 3, 4]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
    
    # Save file
    filename = f"{participant_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = RESULTS_FOLDER / filename
    wb.save(str(filepath))
    
    # Clean up session
    del sessions[session_id]
    
    return jsonify({
        'success': True,
        'filename': filename,
        'filepath': str(filepath)
    })


@app.route('/api/wav-count', methods=['GET'])
def get_wav_count():
    """Get count of WAV files"""
    wav_files = get_wav_files()
    return jsonify({
        'count': len(wav_files),
        'files': wav_files[:10]  # First 10 for verification
    })


if __name__ == '__main__':
    app.run(debug=True, host='localhost', port=5000)
